#!/usr/bin/env python3
"""
PHI Expenditure System - Export Module
Provides CSV, Excel, and QuickBooks IIF export functionality.

Exports Include:
- CSV: Standard comma-separated format for Excel/Google Sheets
- QuickBooks IIF: Import format for QuickBooks Desktop
- Excel: XLSX format with formatting (requires openpyxl)
"""

import csv
import io
from datetime import datetime
from typing import List, Dict, Any


class ExpenditureExporter:
    """Export expenditure data to various formats."""

    @staticmethod
    def to_csv(expenditures: List[Any], include_all_fields: bool = False) -> str:
        """
        Export expenditures to CSV format.

        Args:
            expenditures: List of Expenditure objects
            include_all_fields: If True, include all fields; if False, only key fields

        Returns:
            CSV string
        """
        output = io.StringIO()

        # Define fields to export
        if include_all_fields:
            fieldnames = [
                'id', 'company', 'date', 'vendor', 'vendor_email', 'amount', 'currency',
                'category', 'description', 'invoice_number', 'payment_method',
                'payment_status', 'tax_amount', 'tax_type', 'tax_deductible',
                'confidence', 'verified', 'verified_by', 'verified_at',
                'data_source', 'notes', 'created_at', 'created_by'
            ]
        else:
            fieldnames = [
                'date', 'company', 'vendor', 'description', 'category',
                'amount', 'currency', 'tax_amount', 'invoice_number',
                'payment_method', 'verified'
            ]

        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for exp in expenditures:
            row = {}
            for field in fieldnames:
                value = getattr(exp, field, None)

                # Format dates
                if field in ['date', 'verified_at', 'created_at'] and value:
                    value = value.strftime('%Y-%m-%d') if hasattr(value, 'strftime') else str(value)

                # Format booleans
                elif field in ['verified', 'tax_deductible'] and value is not None:
                    value = 'Yes' if value else 'No'

                # Format decimals
                elif field in ['amount', 'tax_amount'] and value is not None:
                    value = f"{float(value):.2f}"

                row[field] = value or ''

            writer.writerow(row)

        return output.getvalue()


    @staticmethod
    def to_quickbooks_iif(expenditures: List[Any]) -> str:
        """
        Export expenditures to QuickBooks IIF format.

        IIF (Intuit Interchange Format) is used by QuickBooks Desktop.
        Format: Tab-delimited with specific headers.

        Args:
            expenditures: List of Expenditure objects

        Returns:
            IIF formatted string
        """
        output = io.StringIO()

        # IIF Header for transactions
        output.write("!TRNS\tTRNSID\tTRNSTYPE\tDATE\tACCNT\tNAME\tCLASS\tAMOUNT\tDOCNUM\tMEMO\n")
        output.write("!SPL\tSPLID\tTRNSTYPE\tDATE\tACCNT\tNAME\tCLASS\tAMOUNT\tDOCNUM\tMEMO\tCLEAR\n")

        for i, exp in enumerate(expenditures, start=1):
            # Transaction line (TRNS)
            trns_id = f"EXP{i}"
            date = exp.date.strftime('%m/%d/%Y') if hasattr(exp.date, 'strftime') else str(exp.date)
            amount = f"-{float(exp.amount):.2f}"  # Negative for expense
            memo = (exp.description or '')[:100]  # Truncate to 100 chars

            output.write(f"TRNS\t{trns_id}\tCHECK\t{date}\tChecking\t{exp.vendor}\t{exp.company}\t{amount}\t{exp.invoice_number or ''}\t{memo}\n")

            # Split line (SPL) - the expense account
            account = exp.category or "General Expense"
            output.write(f"SPL\t{trns_id}\tCHECK\t{date}\t{account}\t{exp.vendor}\t{exp.company}\t{float(exp.amount):.2f}\t{exp.invoice_number or ''}\t{memo}\tN\n")

            # Tax split line if tax exists
            if exp.tax_amount and float(exp.tax_amount) > 0:
                tax_amount = float(exp.tax_amount)
                output.write(f"SPL\t{trns_id}\tCHECK\t{date}\tSales Tax Payable\t{exp.vendor}\t{exp.company}\t{tax_amount:.2f}\t{exp.invoice_number or ''}\tTax\tN\n")

        output.write("ENDTRNS\n")
        return output.getvalue()


    @staticmethod
    def to_excel(expenditures: List[Any]) -> bytes:
        """
        Export expenditures to Excel XLSX format.

        Requires openpyxl library.

        Args:
            expenditures: List of Expenditure objects

        Returns:
            Excel file as bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Expenditures"

        # Headers
        headers = [
            'Date', 'Company', 'Vendor', 'Description', 'Category',
            'Amount', 'Currency', 'Tax', 'Invoice #', 'Payment Method', 'Verified'
        ]

        # Style header row
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, exp in enumerate(expenditures, start=2):
            ws.cell(row=row_idx, column=1, value=exp.date.strftime('%Y-%m-%d') if hasattr(exp.date, 'strftime') else str(exp.date))
            ws.cell(row=row_idx, column=2, value=exp.company)
            ws.cell(row=row_idx, column=3, value=exp.vendor)
            ws.cell(row=row_idx, column=4, value=exp.description)
            ws.cell(row=row_idx, column=5, value=exp.category)
            ws.cell(row=row_idx, column=6, value=float(exp.amount))
            ws.cell(row=row_idx, column=7, value=exp.currency)
            ws.cell(row=row_idx, column=8, value=float(exp.tax_amount) if exp.tax_amount else 0)
            ws.cell(row=row_idx, column=9, value=exp.invoice_number or '')
            ws.cell(row=row_idx, column=10, value=exp.payment_method)
            ws.cell(row=row_idx, column=11, value='Yes' if exp.verified else 'No')

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


    @staticmethod
    def generate_tax_report_csv(expenditures: List[Any], tax_year: int) -> str:
        """
        Generate CRA-compliant tax report CSV.

        Includes only tax-deductible expenses with required fields.

        Args:
            expenditures: List of Expenditure objects
            tax_year: Tax year (e.g., 2026)

        Returns:
            CSV string
        """
        output = io.StringIO()

        fieldnames = [
            'Date', 'Vendor', 'Description', 'Category', 'Amount (CAD)',
            'Tax Amount (CAD)', 'Tax Type', 'Invoice Number', 'Payment Method'
        ]

        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        total_deductible = 0
        total_tax = 0

        for exp in expenditures:
            # Only include tax-deductible expenses
            if not exp.tax_deductible:
                continue

            # Only include expenses from the specified tax year
            exp_year = exp.date.year if hasattr(exp.date, 'year') else int(str(exp.date)[:4])
            if exp_year != tax_year:
                continue

            amount = float(exp.amount)
            tax = float(exp.tax_amount) if exp.tax_amount else 0

            writer.writerow({
                'Date': exp.date.strftime('%Y-%m-%d') if hasattr(exp.date, 'strftime') else str(exp.date),
                'Vendor': exp.vendor,
                'Description': exp.description[:100],
                'Category': exp.category,
                'Amount (CAD)': f"{amount:.2f}",
                'Tax Amount (CAD)': f"{tax:.2f}",
                'Tax Type': exp.tax_type or '',
                'Invoice Number': exp.invoice_number or '',
                'Payment Method': exp.payment_method
            })

            total_deductible += amount
            total_tax += tax

        # Add summary row
        writer.writerow({})
        writer.writerow({
            'Date': 'TOTAL',
            'Vendor': '',
            'Description': f'Tax Year {tax_year} Summary',
            'Category': '',
            'Amount (CAD)': f"{total_deductible:.2f}",
            'Tax Amount (CAD)': f"{total_tax:.2f}",
            'Tax Type': '',
            'Invoice Number': '',
            'Payment Method': ''
        })

        return output.getvalue()


# Convenience functions for Flask routes
def export_to_csv(expenditures: List[Any], filename: str = None) -> tuple:
    """
    Export to CSV and return Flask response tuple.

    Returns:
        tuple: (csv_data, {'Content-Type': ..., 'Content-Disposition': ...})
    """
    csv_data = ExpenditureExporter.to_csv(expenditures)

    if not filename:
        filename = f"expenditures_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    headers = {
        'Content-Type': 'text/csv',
        'Content-Disposition': f'attachment; filename="{filename}"'
    }

    return csv_data, headers


def export_to_quickbooks(expenditures: List[Any], filename: str = None) -> tuple:
    """
    Export to QuickBooks IIF and return Flask response tuple.

    Returns:
        tuple: (iif_data, {'Content-Type': ..., 'Content-Disposition': ...})
    """
    iif_data = ExpenditureExporter.to_quickbooks_iif(expenditures)

    if not filename:
        filename = f"expenditures_{datetime.now().strftime('%Y%m%d_%H%M%S')}.iif"

    headers = {
        'Content-Type': 'text/plain',
        'Content-Disposition': f'attachment; filename="{filename}"'
    }

    return iif_data, headers


def export_to_excel(expenditures: List[Any], filename: str = None) -> tuple:
    """
    Export to Excel XLSX and return Flask response tuple.

    Returns:
        tuple: (excel_data, {'Content-Type': ..., 'Content-Disposition': ...})
    """
    excel_data = ExpenditureExporter.to_excel(expenditures)

    if not filename:
        filename = f"expenditures_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    headers = {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': f'attachment; filename="{filename}"'
    }

    return excel_data, headers
